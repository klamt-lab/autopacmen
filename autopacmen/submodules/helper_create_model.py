#!/usr/bin/env python3
#
# Copyright 2018-2019 PSB
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""helper_create_model.py

This module contains functions which are useful for a multitide of scripts
which generate AutoPACMEN models.
"""

# IMPORTS
# External modules
import cobra
import copy
import openpyxl
from typing import Any, Dict, List
# Internal modules
from .helper_general import get_float_cell_value


# PRIVATE FUNCTIONS
def _read_stoichiometries_worksheet(workbook: openpyxl.Workbook):
    """Reads a protein stoichiometries worksheet and returns the gene rules and protein stoichiometries.

    This worksheet is a default part of AutoPACMEN's '$PROJECT_NAME_protein_data.xlsx'

    Arguments
    ----------
    * workbook: openpyxl.Workbook ~ An opened XLSX workbook which
      contains an AutoPACMEN 'Stoichiometries of complexes' worksheet.

    Output
    ----------
    Two dictionaries:
    * i) reaction_id_gene_rules_mapping: A mapping of gene rules (as lists) as values
        and reaction IDs as keys. The list is as follows:
        [("COMPLEX_INTERNAL_ID_1", "COMPLEX_INTERNAL_ID_2"), "SINGLE_ID_1", ...]
    * ii) reaction_id_gene_rules_protein_stoichiometry_mapping: A mapping of
         reaction IDs as keys, a tupled form of OR gene rule parts, and
         a list of stoichiometries for the affected gene rule parts.
    """
    # Load the worksheet containing the stoichiometries :D
    worksheet = workbook["Stoichiometries of complexes"]

    # Set mapping variables
    reaction_id_gene_rules_mapping = {}
    reaction_id_gene_rules_protein_stoichiometry_mapping = {}
    for row in worksheet.rows:
        current_cell = 1
        for cell in row:
            if cell.value is None:
                continue
            if current_cell == 1:
                reaction_id = cell.value
                reaction_id_gene_rules_mapping[reaction_id] = []
                reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id] = {
                }
            elif ((current_cell-1) % 2) != 0:
                gene_rule_or_part = str(cell.value)
                if "[" in gene_rule_or_part:
                    gene_rule_or_part = tuple(eval(gene_rule_or_part))
                reaction_id_gene_rules_mapping[reaction_id].append(
                    gene_rule_or_part)
            else:
                stoichiometry = str(cell.value)
                stoichiometries = stoichiometry.split(";")

                i = 0
                for single_stoichiometry in stoichiometries:
                    if len(stoichiometries) > 1:
                        single_protein = gene_rule_or_part[i]
                    else:
                        single_protein = gene_rule_or_part
                    if gene_rule_or_part not in reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id].keys():
                        reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id][gene_rule_or_part] = {
                        }
                    reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id][gene_rule_or_part][single_protein] = \
                        float(single_stoichiometry)
                    i += 1
            current_cell += 1
    return reaction_id_gene_rules_mapping, reaction_id_gene_rules_protein_stoichiometry_mapping


# PUBLIC FUNCTIONS
def add_prot_pool_reaction(model: cobra.Model, id_addition: str,
                           p_total: float, p_measured: float,
                           unmeasured_protein_fraction: float, mean_saturation: float) -> cobra.Model:
    """Adds a protein pool reaction with the given parameters, in accordance with the GECKO paper.

    The protein pool reaction gets the id 'ER_pool'+id_addition

    Arguments
    ----------
    * model: cobra.Model ~ The model to which the protein pool reaction shall be added :D
    * id_addition: str ~ A string that may be added ad the end of the protein pool reaction's id.
    * p_total: float ~ g/gDW of all proteins per 1 gDW cells
    * p_measured: float ~ g/gDW of all proteins with measured concentrations.
    * unmeasured_protein_fraction: float ~ The fraction of the meass of the unmeasured proteins
      on the total protein mass per gDW cells
    * mean_saturation: float ~ A fitted value of all unmeasured protein's mean saturation.

    Output
    ----------
    The given cobra model with a protein pool reaction :D
    """
    # See suppl. data of GECKO paper, equation S28
    pp_reaction = cobra.Reaction("ER_pool"+id_addition)
    pp_reaction.name = "prot_pool reaction for unmeasured proteins"
    pp_reaction.subsystem = "AutoPACMEN"
    pp_reaction.lower_bound = 0
    pp_reaction.upper_bound = (p_total - p_measured) * \
        unmeasured_protein_fraction * mean_saturation
    # The flux is in g/gDW
    prot_pool_metabolite = cobra.Metabolite(
        "prot_pool",
        name="prot_pool pseudometabolite for unmeasured proteins",
        compartment="AutoPACMEN")
    pp_reaction.add_metabolites({prot_pool_metabolite: 1.0})
    model.add_reactions([pp_reaction])
    return model, prot_pool_metabolite


def apply_scenario_on_model(model: cobra.Model, scenario: Dict[str, Any]) -> cobra.Model:
    """Returns a model on which the given scenario is applied.

    Arguments
    ----------
    * model: cobra.Model ~ The model on which the scenario shall be applied.
    * scenario: Dict[str, Any] ~ The scenario's dictionary representation.
    """
    # Change objective if given in the scenario (otherwise, the default obejective is used)
    if "objective" in scenario.keys():
        model.objective = scenario["objective"]
    # Change reactions as given in the scenario
    if "setup" in scenario.keys():
        reactions_to_set_up = scenario["setup"].keys()
        for reaction_to_set_up in reactions_to_set_up:
            reaction_setup = scenario["setup"][reaction_to_set_up]
            if "lower_bound" in reaction_setup.keys():
                new_lower_bound = reaction_setup["lower_bound"]
                model.reactions.get_by_id(
                    reaction_to_set_up).lower_bound = new_lower_bound
            if "upper_bound" in reaction_setup.keys():
                new_upper_bound = reaction_setup["upper_bound"]
                model.reactions.get_by_id(
                    reaction_to_set_up).upper_bound = new_upper_bound
    # Return changed model :D
    return model


def get_irreversible_model(model: cobra.Model, id_addition: str) -> cobra.Model:
    """Returns an irreversible model for further AutoPACMEN processing.

    Only reactions with gene rules, i.e. with proteins, are made irreversible.
    The splitted reactios are then called reaction.id+id_addition+"forward" or
    "reverse", respectively.

    In addition, this function corrects the arm reaction metabolites for
    reversible reactions of sMOMENT models with measured protein concentrations.

    Arguments
    ----------
    * model: cobra.Model ~ The model that shall be made irreversible as described
    * id_addition: str ~ The string that is added before 'reversed' or 'forward'
    """
    model_reaction_ids = [x.id for x in model.reactions]
    for reaction_id in model_reaction_ids:
        reaction = model.reactions.get_by_id(reaction_id)

        if reaction.lower_bound >= 0:
            continue
        if reaction.gene_reaction_rule == "":
            continue

        forward_reaction = copy.deepcopy(reaction)
        forward_reaction.upper_bound = reaction.upper_bound
        forward_reaction.lower_bound = 0
        forward_reaction.id += id_addition + "forward"
        forward_reaction_metabolites_copy = copy.deepcopy(
            forward_reaction.metabolites)
        for key in list(forward_reaction_metabolites_copy.keys()):
            if key.id.startswith("armm_"):
                if key.id.endswith("reverse"):
                    forward_reaction_metabolites_copy[key] = -1
                else:
                    forward_reaction_metabolites_copy[key] = 0
            else:
                forward_reaction_metabolites_copy[key] = 0
        forward_reaction.add_metabolites(forward_reaction_metabolites_copy)
        model.add_reactions([forward_reaction])

        reverse_reaction = copy.deepcopy(reaction)
        reverse_reaction.id += id_addition + "reverse"
        reverse_reaction.upper_bound = -reaction.lower_bound
        reverse_reaction.lower_bound = 0
        reverse_reaction_metabolites_copy = copy.deepcopy(
            reverse_reaction.metabolites)
        for key in list(reverse_reaction_metabolites_copy.keys()):
            if not key.id.startswith("armm_"):
                reverse_reaction_metabolites_copy[key] *= -2
            else:
                if key.id.endswith("reverse"):
                    reverse_reaction_metabolites_copy[key] *= 0
                elif key.id.endswith("forward"):
                    reverse_reaction_metabolites_copy[key] *= -1
        reverse_reaction.add_metabolites(reverse_reaction_metabolites_copy)
        model.add_reactions([reverse_reaction])

        model.remove_reactions([reaction])
    return model


def get_model_with_separated_measured_enzyme_reactions(model: cobra.Model, protein_id_concentration_mapping: Dict[str, float],
                                                       reaction_id_gene_rules_mapping: Dict[str, Any],
                                                       reaction_id_gene_rules_protein_stoichiometry_mapping: Dict[str, Any],
                                                       excluded_reactions: List[str],
                                                       protein_id_mass_mapping: Dict[str, float]):
    """Splits the reactions with measured enzymes in their gene rules according to the OR blocks in the gene rule.

    Arguments
    ----------
    * model: cobra.Model ~ The model for which the reactions shall be splitted.
    * protein_id_concentration_mapping: Dict[str, float] ~ A mapping of protein IDs to measured concentrations.
    * reaction_id_gene_rules_mapping: Dict[str, Any] ~ A mapping of gene rules as list (created with
      _read_stoichiometries_worksheet()) as child with their reaction's IDs as keys
    * reaction_id_gene_rules_protein_stoichiometry_mapping: Dict[str, Any] ~ A dictionary describing the
      internal protein stoichiometries in complexes (created with _read_stoichiometries_worksheet())
    * excluded_reactions: List[str] ~ A list of reactions for which no enzyme constraints shall be added.
    """
    all_measured_proteins = list(protein_id_concentration_mapping.keys())
    print("Measured protein concentration data [mmol/gDW] collected of:")
    print(all_measured_proteins)
    reaction_ids = [x.id for x in model.reactions]
    for reaction_id in reaction_ids:
        reaction = model.reactions.get_by_id(reaction_id)
        if reaction.id not in reaction_id_gene_rules_mapping.keys():
            continue
        if reaction.id in excluded_reactions:
            continue

        gene_rule = reaction_id_gene_rules_mapping[reaction_id]
        # Check if all proteins in the reaction's gene rule have a found mass
        # This is not the case for e.g. spontaneous reactions whcih often get the pseudo-enzyme 's0001'
        all_available = True
        for enzyme in gene_rule:
            if type(enzyme) == str:
                if enzyme not in list(protein_id_mass_mapping.keys()):
                    print(enzyme)
                    all_available = False
                    break
            else:
                for enzyme_id in enzyme:
                    if enzyme_id not in list(protein_id_mass_mapping.keys()):
                        all_available = False
                        break
        # If not all of the mass-checked enzymes have a found mass, ignore this reaction
        if not all_available:
            continue

        # Check for measured proteins in the reaction's gene rule
        # If measured proteins were found, separate for each gene rule element :D
        current_split_no = 1
        measured_elements = []
        unmeasured_elements = []
        gene_rule_as_list = reaction_id_gene_rules_mapping[reaction.id]
        for element in gene_rule_as_list:
            if type(element) is tuple:
                at_least_one_measured = False
                for subelement in element:
                    if subelement in all_measured_proteins:
                        at_least_one_measured = True
                        break
                if at_least_one_measured:
                    measured_elements.append(element)
                else:
                    unmeasured_elements.append(element)
            elif type(element) is str:
                if element in all_measured_proteins:
                    measured_elements.append(element)
                else:
                    unmeasured_elements.append(element)

        if len(measured_elements) == 0:
            continue

        add_arm_reaction = False
        if (len(unmeasured_elements) >= 1) or (len(measured_elements) > 1):
            if reaction.lower_bound >= 0:
                add_arm_reaction = True
                arm_metabolite = cobra.Metabolite(id="armm_"+reaction.id,
                                                  name="arm reaction metabolite for splitting of "+reaction.id,
                                                  compartment="sMOMENT")
                arm_reaction = cobra.Reaction(id="armr_"+reaction.id,
                                              name="Arm reaction for splitting of "+reaction.id,
                                              lower_bound=0,
                                              upper_bound=reaction.upper_bound)
                arm_reaction.add_metabolites({arm_metabolite: -1})
                model.add_reactions([arm_reaction])
            else:
                add_arm_reaction = True
                arm_metabolite_fwd = cobra.Metabolite(id="armm_"+reaction.id+"_forward",
                                                      name="arm reaction metabolite for splitting of "+reaction.id+" forward",
                                                      compartment="sMOMENT")
                arm_reaction_fwd = cobra.Reaction(id="armr_"+reaction.id+"_forward",
                                                  name="Arm reaction for splitting of "+reaction.id,
                                                  lower_bound=0,
                                                  upper_bound=reaction.upper_bound)
                arm_reaction_fwd.add_metabolites({arm_metabolite_fwd: -1})
                model.add_reactions([arm_reaction_fwd])

                arm_metabolite_rev = cobra.Metabolite(id="armm_"+reaction.id+"_reverse",
                                                      name="arm reaction metabolite for splitting of "+reaction.id,
                                                      compartment="sMOMENT")
                arm_reaction_rev = cobra.Reaction(id="armr_"+reaction.id+"_reverse",
                                                  name="Arm reaction for splitting of "+reaction.id,
                                                  lower_bound=0,
                                                  upper_bound=reaction.upper_bound)
                arm_reaction_rev.add_metabolites({arm_metabolite_rev: -1})
                model.add_reactions([arm_reaction_rev])

        # Create reaction for unmeasured elements
        new_reaction = copy.deepcopy(reaction)
        new_reaction.id += "_GPRSPLIT_"+"1"

        new_gene_reaction_rule = ""
        reaction_id_gene_rules_mapping[new_reaction.id] = []
        reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id] = {
        }
        for element in unmeasured_elements:
            reaction_id_gene_rules_mapping[new_reaction.id].append(element)
            if type(element) is tuple:
                for subelement in element:
                    if element not in reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id].keys():
                        reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element] = {
                        }
                    reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element][subelement] = \
                        reaction_id_gene_rules_protein_stoichiometry_mapping[
                            reaction_id][element][subelement]
                if len(new_gene_reaction_rule) > 0:
                    new_gene_reaction_rule += " or "
                new_gene_reaction_rule += "(" + " and ".join(element) + ")"
            else:
                reaction_id_gene_rules_mapping[new_reaction.id].append(element)
                reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element] = {
                }
                reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element][element] = \
                    reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id][element][element]
                if len(new_gene_reaction_rule) > 0:
                    new_gene_reaction_rule += " or "
                new_gene_reaction_rule += element
        model.gene_reaction_rule = new_gene_reaction_rule

        if add_arm_reaction:
            if reaction.lower_bound >= 0:
                new_reaction.add_metabolites({arm_metabolite: 1})
            else:
                new_reaction.add_metabolites(
                    {arm_metabolite_fwd: 1, arm_metabolite_rev: 1})

        if new_gene_reaction_rule != "":
            # print(f"Adding splitted reaction {new_reaction.id}")
            model.add_reactions([new_reaction])
            # print(reaction_id_gene_rules_mapping[new_reaction.id])
            current_split_no = 2
        else:
            current_split_no = 1

        # Create reactions for measured elements
        for element in measured_elements:
            new_reaction = copy.deepcopy(reaction)
            new_reaction.id += "_GPRSPLIT_"+str(current_split_no)
            reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id] = {
            }
            current_split_no += 1
            reaction_id_gene_rules_mapping[new_reaction.id] = [element]
            if type(element) is tuple:
                for subelement in element:
                    if element not in reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id].keys():
                        reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element] = {
                        }
                    reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element][subelement] = \
                        reaction_id_gene_rules_protein_stoichiometry_mapping[
                            reaction_id][element][subelement]
                new_reaction.gene_reaction_rule = " and ".join(element)
            else:
                reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element] = {
                }
                reaction_id_gene_rules_protein_stoichiometry_mapping[new_reaction.id][element][element] = \
                    reaction_id_gene_rules_protein_stoichiometry_mapping[reaction_id][element][element]
                new_reaction.gene_reaction_rule = element

            # print("Adding new gene-rule-splitted reaction", new_reaction.id)
            if add_arm_reaction:
                if reaction.lower_bound >= 0:
                    new_reaction.add_metabolites({arm_metabolite: 1})
                else:
                    new_reaction.add_metabolites(
                        {arm_metabolite_fwd: 1, arm_metabolite_rev: 1})
            model.add_reactions([new_reaction])
            # print(reaction_id_gene_rules_mapping[new_reaction.id])
        model.remove_reactions([reaction])

    return model, reaction_id_gene_rules_mapping, reaction_id_gene_rules_protein_stoichiometry_mapping


def get_p_measured(protein_id_concentration_mapping: Dict[str, float],
                   protein_id_mass_mapping: Dict[str, float]) -> float:
    """Return p_measured (as defined in the GECKO paper) using the given arguments.

    Arguments
    ----------
    * protein_id_concentration_mapping: Dict[str, float] ~ The given concentrations
      of the proteins.
    * protein_id_mass_mapping: Dict[str, float] ~ The masses of the given proteins.
    """
    p_measured = .0
    measured_protein_ids = (protein_id_concentration_mapping.keys())
    for measured_protein_id in measured_protein_ids:
        # mmol/gDW
        protein_concentration = protein_id_concentration_mapping[measured_protein_id]
        # Da = g/mol
        protein_mass = protein_id_mass_mapping[measured_protein_id]
        # (mol/gDW) * (g/mol) = g/gDW
        protein_mass_concentration = (
            protein_concentration/1000) * protein_mass
        p_measured += protein_mass_concentration
    return p_measured


def read_enzyme_stoichiometries_xlsx(basepath: str):
    """Reads the internal protein stoichiometries for each enzyme XLSX.

    Arguments
    ----------
    * basepath: str ~ The path in which the XLSX can be found.
    """
    workbook = openpyxl.load_workbook(
        filename=basepath+"_enzyme_stoichiometries.xlsx", read_only=True)
    reaction_id_gene_rules_mapping, reaction_id_gene_rules_protein_stoichiometry_mapping = \
        _read_stoichiometries_worksheet(workbook)

    return reaction_id_gene_rules_mapping, reaction_id_gene_rules_protein_stoichiometry_mapping


def read_protein_data_xlsx(basepath: str):
    """Reads the protein data XLSX '$PROJECT_NAME_protein_data.xlsx'

    Argument
    ----------
    * basepath: str ~ The folder in which the XLSX exists.

    Output
    ----------
    4 values:
    1. protein_id_concentration_mapping: Dict[str, float] ~ A dictionary with each protein ID
       as key the associated protein concentration as child.
    2. p_total: float ~ The full fraction of proteins per gram biomass (g/gDW)
    3. unmeasured_protein_fraction: float ~ The mass fraction of the proteins without a measured
       concentration.
    4. mean_saturation: float ~ A given - and usually fitted - value of mean saturation for all
       enzymes.
    """
    protein_id_concentration_mapping: Dict[str, float] = {}
    workbook = openpyxl.load_workbook(
        filename=basepath+"_protein_data.xlsx", read_only=True)

    worksheet = workbook["Total protein data"]
    p_total_value = worksheet.cell(row=1, column=2).value
    unmeasured_protein_fraction_value = worksheet.cell(row=2, column=2).value
    mean_saturation_value = worksheet.cell(row=3, column=2).value
    p_total = get_float_cell_value(p_total_value)
    unmeasured_protein_fraction = get_float_cell_value(
        unmeasured_protein_fraction_value)
    mean_saturation = get_float_cell_value(mean_saturation_value)

    worksheet2 = workbook["Single protein data"]
    row = 2
    while True:
        protein_id = worksheet2.cell(row=row, column=1).value
        if protein_id is None:
            break
        protein_concentration_value = worksheet2.cell(row=row, column=2).value
        protein_concentration = get_float_cell_value(
            protein_concentration_value)
        protein_id_concentration_mapping[protein_id] = protein_concentration
        row += 1

    return protein_id_concentration_mapping, p_total, unmeasured_protein_fraction, mean_saturation
